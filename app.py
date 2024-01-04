import openpyxl
import os
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import Calendar
from datetime import datetime #timedelta
import tkinter.simpledialog as simpledialog
from openpyxl.styles import  PatternFill, Font

FERIADOS = ["25/12/2023", "01/01/2024", "31/09/2023"]  # Exemplo de feriados.

def create_schedule():
    try:
        if os.path.exists("HorarioFuncionarios.xlsx"):
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 'Funcionário'
        wb.save("HorarioFuncionarios.xlsx")
    except Exception as e:
        print(f"Erro ao criar a planilha: {e}")

class ScheduleManager:
    def __init__(self, filename):
        self.filename = filename  # Certifique-se de que esta linha esteja presente
        try:
            self.wb = openpyxl.load_workbook(filename)
            self.ws = self.wb.active
        except Exception as e:
            print(f"Erro ao carregar a planilha: {e}")



    def fill_schedule(self):
        for row, employee_cell in enumerate(self.ws['A'], start=1):
            if row == 1:  # Skip header
                continue
            employee_name = employee_cell.value
            if not employee_name:  # If no more employees, stop
                break

            print(f"Registrar horários para {employee_name}:")
            for col, day in enumerate(['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo'], start=2):
                time = input(f"Horário para {day}: ")
                self.ws.cell(row=row, column=col, value=time)

    def check_week_off(self):
        for row_num, row in enumerate(self.ws.iter_rows(min_row=2, max_col=9), start=2):
            took_off = any(cell.value and "folga" in cell.value.lower() for cell in row[1:8])
            status_cell = self.ws.cell(row=row_num, column=9)  # Assuming 9th column is Status

            if status_cell.value == "Took Off" and took_off:
                status_cell.value = "Cannot Work Home Office"
            elif took_off:
                status_cell.value = "Took Off"
            else:
                status_cell.value = ""
    
    def add_new_employee(self, name):
        last_row = self.ws.max_row
        self.ws.cell(row=last_row + 1, column=1, value=name)

    def get_employee_names(self):
        """Retorna uma lista com os nomes dos funcionários da planilha."""
        return [cell.value for cell in self.ws['A'] if cell.value and cell.row != 1]

    def register_date(self, name, date, option):
        try:
            # Verifica se o nome do funcionário está na planilha
            row_num = next((row for row, cell in enumerate(self.ws['A']) if cell.value == name), None)
            if row_num is None:
                raise ValueError(f"Funcionário '{name}' não encontrado.")

            # Incrementa para ajustar o índice da linha (começa em 1, não em 0)
            row_num += 1

            # Obtém o número da coluna para a data
            col_num = self.get_date_column(date)

            # Se encontrou o número da linha e da coluna, procede com a atualização
            if row_num and col_num:
                cell = self.ws.cell(row=row_num, column=col_num, value=option)
                
                # Altera a cor de fundo da célula se a opção for "Folga"
                if option == "Folga":
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                else:
                    cell.fill = PatternFill(None)  # Reverte para o padrão se a opção não for "Folga"
                
                # Salva as alterações na planilha
                self.save()
        except ValueError as ve:
            print(ve)
        except Exception as e:
            print(f"Erro ao registrar a data: {e}")

    def check_right_for_homeoffice(self, name, date):
        dt = datetime.strptime(date, '%d/%m/%Y')
        if dt.weekday() >= 5 or date in FERIADOS:
            return True
        return False
    
    def get_date_column(self, date):
        """Retorna o número da coluna para uma data específica, criando-a se não existir."""
        for col_num, col_cells in enumerate(self.ws.iter_cols(min_row=1, max_row=1)):
            if col_cells[0].value == date:
                return col_num + 1
        # Se não encontrou a coluna, crie uma nova
        last_col = self.ws.max_column + 1
        self.ws.cell(row=1, column=last_col, value=date)
        return last_col

    def register_date(self, name, date, option):
        row_num = next((row for row, cell in enumerate(self.ws['A']) if cell.value == name), None) + 1
        col_num = self.get_date_column(date)
        if row_num and col_num:
            self.ws.cell(row=row_num, column=col_num, value=option)
            self.save()

    def check_right_for_homeoffice(self, name, date):
        dt = datetime.strptime(date, '%d/%m/%Y')
        if dt.weekday() >= 5 or date in FERIADOS:
            return True
        return False
    
    def delete_employee(self, name):
        for row_num, cell in enumerate(self.ws['A'], start=1):
            if cell.value == name:
                self.ws.delete_rows(row_num)
                self.save()
                return True
        return False
    
    def get_all_employees(self):
        return [cell.value for cell in self.ws['A'] if cell.value and cell.row != 1]

    def save(self):
        try:
            self.wb.save(self.filename)
        except Exception as e:
            print(f"Erro ao salvar a planilha: {e}")

class App:
    def __init__(self, root, manager):
        self.root = root
        self.root.title("Gerenciador de Horários")
        self.manager = manager

        self.calendar = Calendar(self.root, selectmode='day', date_pattern='dd/mm/yyyy')
        self.calendar.pack(pady=20)

        self.employee_label = ttk.Label(self.root, text="Funcionário:")
        self.employee_label.pack(pady=5)

        self.employee_dropdown = ttk.Combobox(self.root, values=self.manager.get_employee_names())
        self.employee_dropdown.pack(pady=5)

        self.option_var = tk.StringVar()
        self.folga_radio = ttk.Radiobutton(self.root, text="Folga", variable=self.option_var, value="Folga")
        self.folga_radio.pack(pady=5)
        
        self.t1_radio = ttk.Radiobutton(self.root, text="T1 (Presencial)", variable=self.option_var, value="T1")
        self.t1_radio.pack(pady=5)
        
        self.t2_radio = ttk.Radiobutton(self.root, text="T2 (Home Office)", variable=self.option_var, value="T2")
        self.t2_radio.pack(pady=5)

        self.register_btn = ttk.Button(self.root, text="Registrar", command=self.register_schedule)
        self.register_btn.pack(pady=20)

        self.add_employee_btn = ttk.Button(self.root, text="Adicionar Funcionário", command=self.add_employee)
        self.add_employee_btn.pack(pady=20)


    def register_schedule(self):
        selected_date = datetime.strptime(self.calendar.get_date(), '%d/%m/%Y')
        current_date = datetime.now()

        if selected_date > current_date:
            messagebox.showwarning("Aviso", "Não é permitido registrar datas futuras!")
            return

        name = self.employee_dropdown.get()
        option = self.option_var.get()

        # Marcar feriados automaticamente como "Folga"
        if self.calendar.get_date() in FERIADOS:
            option = "Folga"

        self.manager.register_date(name, self.calendar.get_date(), option)
        messagebox.showinfo("Aviso", "Registrado com sucesso!")
        
        if self.manager.check_right_for_homeoffice(name, self.calendar.get_date()):
            messagebox.showinfo("Aviso", f"{name} tem direito a T2 (Home Office) na próxima semana!")

    def add_employee(self):
        name = tk.simpledialog.askstring("Nome do Funcionário", "Insira o nome do novo funcionário:")
        if name:
            self.manager.add_new_employee(name)
            self.employee_dropdown["values"] = self.manager.get_employee_names()
            messagebox.showinfo("Sucesso", "Funcionário cadastrado com sucesso!")

    def delete_employee(self):
        name = self.employee_dropdown.get()
        if not name:
            messagebox.showerror("Erro", "Selecione um funcionário para deletar!")
            return
    
        confirm = messagebox.askyesno("Confirmar", f"Tem certeza de que deseja deletar {name}?")
        if confirm:
            deleted = self.manager.delete_employee(name)
            if deleted:
                messagebox.showinfo("Sucesso", f"{name} foi deletado com sucesso!")
                self.employee_dropdown["values"] = self.manager.get_all_employees()
                self.employee_dropdown.set('')
            else:
                messagebox.showerror("Erro", f"Não foi possível deletar {name}!")


if __name__ == "__main__":
    create_schedule()  # Criar a planilha, se ela não existir.
    file_name = "HorarioFuncionarios.xlsx"
    manager = ScheduleManager(file_name)
    root = tk.Tk()
    app = App(root, manager)
    root.mainloop()

